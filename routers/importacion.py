# -*- coding: utf-8 -*-
"""
ALTA MASIVA POR EXCEL — clientes y personal
===========================================
Para arrancar el sistema con toda la cartera de un jalón, en vez de capturar
cliente por cliente.

CÓMO FUNCIONA (tres pasos, sin sorpresas):
  1. El Director descarga la PLANTILLA (GET /api/importacion/plantilla).
     Trae dos hojas —CLIENTES y PERSONAL—, una fila de ejemplo, y una hoja de
     INSTRUCCIONES con la lista exacta de regímenes válidos.
  2. La llena y la sube en modo REVISIÓN (POST /api/importacion/revisar).
     El sistema valida TODO sin guardar nada y devuelve, renglón por renglón,
     qué está bien y qué está mal (con el motivo). Nada entra a medias.
  3. Si la revisión está limpia, CONFIRMA (POST /api/importacion/confirmar).
     Ahí sí se dan de alta, y se devuelven las contraseñas temporales de los
     clientes a los que se les creó portal, para entregárselas.

REGLA: si un solo renglón tiene un error, la importación NO se ejecuta. Es
preferible corregir el Excel a quedarse con media cartera cargada.
"""
import io
import re
from datetime import datetime

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from database import get_db
from models.models import Cliente, RolUsuario, TipoCliente, Usuario
from services import auditoria, fiscal_v2
from services.auth import hash_password, solo_director

router = APIRouter(prefix="/api/importacion", tags=["Alta masiva"])

# --- Columnas de la plantilla (el orden es el que ve el usuario) ---
COLUMNAS_CLIENTES = [
    ("nombre_comercial", "Nombre comercial *", "Ferretería El Tornillo"),
    ("razon_social", "Razón social *", "El Tornillo SA de CV"),
    ("rfc", "RFC *", "TOR990101AB1"),
    ("telefono_whatsapp", "WhatsApp *", "+524921234567"),
    ("email", "Correo", "contacto@eltornillo.mx"),
    ("tipo_persona", "Tipo de persona *", "moral"),
    ("regimen_fiscal", "Régimen fiscal *", "pm_general"),
    ("tipo_cliente", "Tipo de cliente", "estandar"),
    ("contador_email", "Correo del contador asignado", "carlos@pya.mx"),
    ("tiene_imss", "¿Tiene IMSS? (si/no)", "si"),
    ("tiene_nomina", "¿Tiene nómina? (si/no)", "si"),
    ("periodicidad_nomina", "Periodicidad de nómina", "quincenal"),
    # COBRANZA: de aquí come el módulo de cobranza. Sin esto queda de adorno.
    ("honorario_mensual", "Honorario ($) *", "3500"),
    ("periodicidad_honorario", "Periodicidad del honorario", "mensual"),
    ("dia_corte_honorario", "Día de corte (1-28)", "1"),
    ("adeudo_previo_monto", "Adeudo anterior ($)", "0"),
    ("adeudo_previo_concepto", "Concepto del adeudo anterior", "Honorarios 2025 pendientes"),
    # Bases de CONTPAQi (opcionales; el agente local las usa por RFC)
    ("bd_contpaq_contabilidad", "Base CONTPAQ contabilidad", "ctELTORNILLO"),
    ("bd_contpaq_nomina", "Base CONTPAQ nóminas", "nomELTORNILLO"),
    ("crear_portal", "¿Crear su portal? (si/no)", "si"),
]

COLUMNAS_PERSONAL = [
    ("nombre", "Nombre completo *", "C.P. Ana Ramírez"),
    ("email", "Correo *", "ana@pya.mx"),
    ("rol", "Rol *", "contador"),
]

ROLES_VALIDOS = {
    "administrador": RolUsuario.ADMINISTRADOR,
    "director": RolUsuario.DIRECTOR,
    "supervisor": RolUsuario.SUPERVISOR,
    "admin_secretaria": RolUsuario.ADMIN_SECRETARIA,
    "contador": RolUsuario.CONTADOR,
}
TIPOS_CLIENTE = {t.value: t for t in TipoCliente}
PERIODICIDADES = ("semanal", "quincenal", "mensual")
RFC_PATRON = re.compile(r"^[A-ZÑ&]{3,4}\d{6}[A-Z0-9]{3}$")

AZUL = "0A5AA1"
AMARILLO = "FFF3CD"
GRIS = "F2F4F7"


# ===========================================================================
# 1) PLANTILLA
# ===========================================================================

def _encabezar(hoja, columnas):
    for i, (_, etiqueta, ejemplo) in enumerate(columnas, start=1):
        celda = hoja.cell(row=1, column=i, value=etiqueta)
        celda.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        celda.fill = PatternFill("solid", fgColor=AZUL)
        celda.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        # Fila 2: EJEMPLO (se borra antes de llenar; el sistema la ignora)
        ej = hoja.cell(row=2, column=i, value=ejemplo)
        ej.font = Font(name="Arial", size=10, italic=True, color="8A93A6")
        ej.fill = PatternFill("solid", fgColor=AMARILLO)
        hoja.column_dimensions[get_column_letter(i)].width = max(
            16, min(30, len(etiqueta) + 4))
    hoja.row_dimensions[1].height = 32
    hoja.freeze_panes = "A3"


@router.get("/plantilla")
def descargar_plantilla(u: Usuario = Depends(solo_director)):
    """
    Genera la plantilla de Excel. Trae la lista EXACTA de regímenes válidos
    para que nadie tenga que adivinar cómo se escriben.
    """
    wb = Workbook()

    # --- Hoja 1: INSTRUCCIONES ---
    ins = wb.active
    ins.title = "INSTRUCCIONES"
    ins.column_dimensions["A"].width = 34
    ins.column_dimensions["B"].width = 62

    def linea(fila, a, b="", negrita=False, titulo=False):
        ca = ins.cell(row=fila, column=1, value=a)
        cb = ins.cell(row=fila, column=2, value=b)
        ca.font = Font(name="Arial", size=11 if titulo else 10,
                       bold=negrita or titulo,
                       color=AZUL if titulo else "111826")
        cb.font = Font(name="Arial", size=10)
        cb.alignment = Alignment(wrap_text=True, vertical="top")

    linea(1, "ALTA MASIVA — P&A Despacho Contable", "", titulo=True)
    linea(3, "Cómo se usa", "", negrita=True)
    linea(4, "1.", "Llene las hojas CLIENTES y PERSONAL. Puede llenar solo una.")
    linea(5, "2.", "BORRE la fila del ejemplo (la amarilla). Si la deja, se ignora.")
    linea(6, "3.", "Suba el archivo en Administración → Alta masiva. Primero verá "
                   "una REVISIÓN: le dirá renglón por renglón qué está bien y qué no.")
    linea(7, "4.", "Si todo está correcto, confirme. Si hay UN solo error, no se "
                   "da de alta nada: corrija el Excel y vuelva a subirlo.")
    linea(9, "Notas importantes", "", negrita=True)
    linea(10, "Columnas con *", "Son obligatorias.")
    linea(11, "RFC", "Sin espacios ni guiones. Ej. TOR990101AB1")
    linea(12, "Régimen fiscal", "Copie la clave EXACTA de la lista de abajo. "
                                "Sin régimen, la calculadora no funciona para ese cliente.")
    linea(13, "¿Crear su portal?", "Si pone 'si', se genera una contraseña temporal "
                                   "que se le mostrará al terminar, para entregársela "
                                   "al cliente. Requiere que el cliente tenga correo.")
    linea(14, "Honorario *", "Lo que se le cobra al cliente por periodo. Es "
                             "OBLIGATORIO: de aquí come el módulo de cobranza. "
                             "Escriba solo el número (3500), sin signo de pesos.")
    linea(15, "Periodicidad del honorario", "mensual, bimestral o anual.")
    linea(16, "Día de corte", "Día del mes en que se genera el cobro (1 a 28).")
    linea(17, "Adeudo anterior", "Lo que el cliente YA debía antes de entrar al "
                                 "sistema. Déjelo en 0 si no debe nada. Aparecerá "
                                 "en su estado de cuenta como saldo anterior.")
    linea(18, "Bases CONTPAQ", "Opcionales. Los nombres de las bases de datos de "
                               "CONTPAQi de ese cliente (el agente local las usa "
                               "para traer sus resultados). Ej. ctELTORNILLO")
    linea(14, "Contador asignado", "Escriba el CORREO de quien lleva su contabilidad "
                                   "(debe existir ya, o venir en la hoja PERSONAL).")
    linea(15, "Personal", "A cada persona se le genera una contraseña temporal; "
                          "deberá cambiarla en su primer acceso.")

    fila = 17
    linea(fila, "REGÍMENES VÁLIDOS (copie la clave)", "", negrita=True)
    fila += 1
    ins.cell(row=fila, column=1, value="CLAVE").font = Font(name="Arial", bold=True, size=9)
    ins.cell(row=fila, column=2, value="RÉGIMEN").font = Font(name="Arial", bold=True, size=9)
    fila += 1
    for grupo, claves in fiscal_v2.GRUPOS_REGIMEN.items():
        if grupo == "Declaraciones anuales":
            continue  # las anuales no son "el régimen del cliente"
        c = ins.cell(row=fila, column=1, value=f"— {grupo} —")
        c.font = Font(name="Arial", bold=True, size=9, color=AZUL)
        fila += 1
        for clave in claves:
            ins.cell(row=fila, column=1, value=clave).font = Font(name="Courier New", size=9)
            ins.cell(row=fila, column=2, value=fiscal_v2.REGIMENES[clave]).font = Font(name="Arial", size=9)
            fila += 1

    fila += 1
    linea(fila, "ROLES VÁLIDOS (hoja PERSONAL)", "", negrita=True)
    fila += 1
    for clave, desc in [
        ("contador", "Elabora cálculos. NUNCA autoriza los suyos."),
        ("admin_secretaria", "Cobranza y contabilidades. No autoriza."),
        ("supervisor", "Contadora y autorizadora (puede autofirmar)."),
        ("director", "Acceso total al despacho."),
        ("administrador", "Acceso técnico total al sistema."),
    ]:
        ins.cell(row=fila, column=1, value=clave).font = Font(name="Courier New", size=9)
        ins.cell(row=fila, column=2, value=desc).font = Font(name="Arial", size=9)
        fila += 1

    # --- Hoja 2: CLIENTES ---
    hc = wb.create_sheet("CLIENTES")
    _encabezar(hc, COLUMNAS_CLIENTES)

    # --- Hoja 3: PERSONAL ---
    hp = wb.create_sheet("PERSONAL")
    _encabezar(hp, COLUMNAS_PERSONAL)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = f"alta_masiva_pafirma_{datetime.now():%Y%m%d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'})


# ===========================================================================
# 2) LECTURA Y VALIDACIÓN
# ===========================================================================

def _txt(v) -> str:
    return str(v).strip() if v is not None else ""


def _si_no(v) -> bool:
    return _txt(v).lower() in ("si", "sí", "s", "x", "1", "true", "verdadero")


def _es_ejemplo(fila: dict, columnas, renglon: int) -> bool:
    """
    La fila amarilla de ejemplo se ignora si no la borraron. Solo se busca en
    el RENGLÓN 2 (donde la puso la plantilla): si se comparara por contenido
    en cualquier renglón, un cliente real que coincidiera con el ejemplo se
    perdería en silencio.
    """
    if renglon != 2:
        return False
    ejemplos = {c[0]: c[2] for c in columnas}
    llenos = [k for k, v in fila.items() if v]
    return bool(llenos) and all(fila.get(k, "") == ejemplos.get(k, "")
                                for k in llenos)


def _leer_hoja(wb, nombre, columnas):
    if nombre not in wb.sheetnames:
        return []
    hoja = wb[nombre]
    filas = []
    for i, valores in enumerate(hoja.iter_rows(min_row=2, values_only=True),
                                start=2):
        fila = {c[0]: _txt(v) for c, v in zip(columnas, valores)}
        if not any(fila.values()):
            continue                     # renglón vacío
        if _es_ejemplo(fila, columnas, i):
            continue                     # la fila de ejemplo (solo renglón 2)
        fila["_renglon"] = i
        filas.append(fila)
    return filas


def _numero(valor):
    """Lee un número del Excel aunque venga como '$3,500.00' o con espacios."""
    if valor is None or str(valor).strip() == "":
        return None
    try:
        limpio = str(valor).replace("$", "").replace(",", "").replace(" ", "")
        return round(float(limpio), 2)
    except (TypeError, ValueError):
        return None


def _validar_clientes(filas, db, correos_personal_nuevo):
    resultados = []
    rfcs_vistos, correos_vistos = set(), set()
    for f in filas:
        errores = []
        for campo, etiqueta in [("nombre_comercial", "nombre comercial"),
                                ("razon_social", "razón social"),
                                ("rfc", "RFC"),
                                ("telefono_whatsapp", "WhatsApp")]:
            if not f[campo]:
                errores.append(f"falta {etiqueta}")

        rfc = f["rfc"].upper().replace(" ", "").replace("-", "")
        if rfc and not RFC_PATRON.match(rfc):
            errores.append(f"RFC con formato inválido ({rfc})")
        if rfc in rfcs_vistos:
            errores.append(f"RFC repetido en el archivo ({rfc})")
        rfcs_vistos.add(rfc)

        tp = f["tipo_persona"].lower()
        if tp not in ("fisica", "física", "moral"):
            errores.append("tipo de persona debe ser 'fisica' o 'moral'")
        tp = "fisica" if tp.startswith("f") else "moral"

        reg = f["regimen_fiscal"].lower().strip()
        if not reg:
            errores.append("falta el régimen fiscal (sin él, la calculadora "
                           "no funciona para este cliente)")
        elif reg not in fiscal_v2.REGIMENES or reg.startswith("anual_"):
            errores.append(f"régimen no válido: '{reg}' (vea la lista en la "
                           f"hoja INSTRUCCIONES)")
        elif fiscal_v2.TIPO_PERSONA_DE_REGIMEN.get(reg) != tp:
            correcto = fiscal_v2.TIPO_PERSONA_DE_REGIMEN.get(reg)
            errores.append(f"el régimen '{reg}' es de persona {correcto}, "
                           f"pero capturó '{tp}'")

        tc = f["tipo_cliente"].lower() or "estandar"
        if tc not in TIPOS_CLIENTE:
            errores.append(f"tipo de cliente no válido: '{tc}' "
                           f"({', '.join(TIPOS_CLIENTE)})")

        per = f["periodicidad_nomina"].lower() or "quincenal"
        if per not in PERIODICIDADES:
            errores.append(f"periodicidad no válida: '{per}'")

        correo = f["email"].lower()
        if correo:
            if "@" not in correo:
                errores.append(f"correo inválido: {correo}")
            if correo in correos_vistos:
                errores.append(f"correo repetido en el archivo: {correo}")
            correos_vistos.add(correo)
            if db.query(Usuario).filter(Usuario.email == correo).first():
                errores.append(f"ya existe un usuario con el correo {correo}")

        if _si_no(f["crear_portal"]) and not correo:
            errores.append("para crear el portal se necesita el correo del cliente")

        cont = f["contador_email"].lower()
        contador_id = None
        if cont:
            usuario = db.query(Usuario).filter(Usuario.email == cont).first()
            if usuario:
                if usuario.rol == RolUsuario.CLIENTE:
                    errores.append(f"{cont} es un cliente, no personal")
                contador_id = usuario.id
            elif cont not in correos_personal_nuevo:
                errores.append(f"no existe personal con el correo {cont} "
                               f"(ni viene en la hoja PERSONAL)")

        # ¿RFC ya dado de alta? (comparación en memoria: el RFC va cifrado)
        if rfc and any(c.rfc == rfc for c in db.query(Cliente).all()):
            errores.append(f"ya existe un cliente con el RFC {rfc}")

        # COBRANZA: el honorario es obligatorio; sin él la cobranza no arranca
        honorario = _numero(f.get("honorario_mensual"))
        if honorario is None or honorario <= 0:
            errores.append("falta el honorario (es el que cobra la cobranza)")
        per_hon = (f.get("periodicidad_honorario") or "mensual").strip().lower()
        if per_hon not in ("mensual", "bimestral", "anual"):
            errores.append(f"periodicidad del honorario '{per_hon}' no válida "
                           f"(mensual, bimestral o anual)")
            per_hon = "mensual"
        dia_corte = int(_numero(f.get("dia_corte_honorario")) or 1)
        if not 1 <= dia_corte <= 28:
            errores.append("el día de corte debe estar entre 1 y 28")
            dia_corte = 1
        adeudo = _numero(f.get("adeudo_previo_monto")) or 0
        if adeudo < 0:
            errores.append("el adeudo anterior no puede ser negativo")
            adeudo = 0

        resultados.append({
            "renglon": f["_renglon"], "nombre": f["nombre_comercial"],
            "errores": errores, "ok": not errores,
            "_datos": {
                "nombre_comercial": f["nombre_comercial"],
                "razon_social": f["razon_social"], "rfc": rfc,
                "telefono_whatsapp": f["telefono_whatsapp"],
                "email": correo or None, "tipo_persona": tp,
                "regimen_fiscal": reg, "tipo_cliente": tc,
                "contador_email": cont or None, "contador_id": contador_id,
                "tiene_imss": _si_no(f["tiene_imss"]),
                "tiene_nomina": _si_no(f["tiene_nomina"]),
                "periodicidad_nomina": per,
                "honorario_mensual": honorario,
                "periodicidad_honorario": per_hon,
                "dia_corte_honorario": dia_corte,
                "adeudo_previo_monto": adeudo,
                "adeudo_previo_concepto": (f.get("adeudo_previo_concepto") or "").strip() or None,
                "bd_contpaq_contabilidad": (f.get("bd_contpaq_contabilidad") or "").strip() or None,
                "bd_contpaq_nomina": (f.get("bd_contpaq_nomina") or "").strip() or None,
                "crear_portal": _si_no(f["crear_portal"]),
            }})
    return resultados


def _validar_personal(filas, db):
    resultados = []
    correos_vistos = set()
    for f in filas:
        errores = []
        if not f["nombre"]:
            errores.append("falta el nombre")
        correo = f["email"].lower()
        if not correo:
            errores.append("falta el correo")
        elif "@" not in correo:
            errores.append(f"correo inválido: {correo}")
        elif correo in correos_vistos:
            errores.append(f"correo repetido en el archivo: {correo}")
        elif db.query(Usuario).filter(Usuario.email == correo).first():
            errores.append(f"ya existe un usuario con el correo {correo}")
        correos_vistos.add(correo)

        rol = f["rol"].lower()
        if rol not in ROLES_VALIDOS:
            errores.append(f"rol no válido: '{rol}' "
                           f"({', '.join(ROLES_VALIDOS)})")

        resultados.append({
            "renglon": f["_renglon"], "nombre": f["nombre"],
            "errores": errores, "ok": not errores,
            "_datos": {"nombre": f["nombre"], "email": correo, "rol": rol}})
    return resultados


def _procesar(archivo: UploadFile, db: Session):
    if not (archivo.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "El archivo debe ser un Excel (.xlsx)")
    try:
        wb = load_workbook(io.BytesIO(archivo.file.read()), data_only=True)
    except Exception:
        raise HTTPException(400, "No se pudo leer el archivo. Use la plantilla "
                                 "descargada del sistema.")
    if "CLIENTES" not in wb.sheetnames and "PERSONAL" not in wb.sheetnames:
        raise HTTPException(400, "El archivo no tiene las hojas CLIENTES ni "
                                 "PERSONAL. Descargue la plantilla del sistema.")

    filas_personal = _leer_hoja(wb, "PERSONAL", COLUMNAS_PERSONAL)
    filas_clientes = _leer_hoja(wb, "CLIENTES", COLUMNAS_CLIENTES)
    personal = _validar_personal(filas_personal, db)
    correos_nuevos = {p["_datos"]["email"] for p in personal if p["ok"]}
    clientes = _validar_clientes(filas_clientes, db, correos_nuevos)
    return personal, clientes


def _resumen(personal, clientes):
    errores = ([{"hoja": "PERSONAL", **p} for p in personal if not p["ok"]]
               + [{"hoja": "CLIENTES", **c} for c in clientes if not c["ok"]])
    for e in errores:
        e.pop("_datos", None)
    return {
        "personal": {"total": len(personal),
                     "listos": sum(1 for p in personal if p["ok"])},
        "clientes": {"total": len(clientes),
                     "listos": sum(1 for c in clientes if c["ok"])},
        "errores": errores,
        # bool() explícito: `and` en Python devuelve el último operando, no
        # un booleano —esto llegaba al frontend como una lista.
        "se_puede_importar": bool(not errores and (personal or clientes)),
    }


@router.post("/revisar")
def revisar(request: Request, archivo: UploadFile = File(...),
            u: Usuario = Depends(solo_director),
            db: Session = Depends(get_db)):
    """
    REVISIÓN sin guardar nada. Devuelve renglón por renglón qué está bien y
    qué está mal, con el motivo exacto. Ni un registro entra a la base aquí.
    """
    personal, clientes = _procesar(archivo, db)
    res = _resumen(personal, clientes)
    res["vista_previa"] = {
        "personal": [{"renglon": p["renglon"], "nombre": p["_datos"]["nombre"],
                      "correo": p["_datos"]["email"], "rol": p["_datos"]["rol"]}
                     for p in personal if p["ok"]][:50],
        "clientes": [{"renglon": c["renglon"],
                      "nombre": c["_datos"]["nombre_comercial"],
                      "regimen": fiscal_v2.REGIMENES.get(c["_datos"]["regimen_fiscal"], "—"),
                      "portal": c["_datos"]["crear_portal"]}
                     for c in clientes if c["ok"]][:50],
    }
    return res


@router.post("/confirmar")
def confirmar(request: Request, archivo: UploadFile = File(...),
              u: Usuario = Depends(solo_director),
              db: Session = Depends(get_db)):
    """
    Da de alta TODO lo del archivo. Si hay un solo error, no se guarda nada
    (todo o nada: es preferible corregir el Excel a quedar a medias).
    Devuelve las CONTRASEÑAS TEMPORALES generadas, para entregarlas.
    """
    from routers.admin import generar_password_temporal

    personal, clientes = _procesar(archivo, db)
    res = _resumen(personal, clientes)
    if not res["se_puede_importar"]:
        raise HTTPException(400, {
            "mensaje": "El archivo tiene errores: no se dio de alta nada. "
                       "Corrija los renglones señalados y vuelva a subirlo.",
            "errores": res["errores"]})

    credenciales = []

    # 1) PERSONAL primero (los clientes pueden referirse a ellos)
    creados_personal = {}
    for p in personal:
        d = p["_datos"]
        temporal = generar_password_temporal()
        usuario = Usuario(nombre=d["nombre"], email=d["email"],
                          rol=ROLES_VALIDOS[d["rol"]],
                          password_hash=hash_password(temporal),
                          debe_cambiar_password=True)
        db.add(usuario)
        db.flush()
        creados_personal[d["email"]] = usuario.id
        credenciales.append({"tipo": "Personal", "nombre": d["nombre"],
                             "usuario": d["email"], "password": temporal,
                             "rol": d["rol"]})

    # 2) CLIENTES
    for cfila in clientes:
        d = cfila["_datos"]
        contador_id = d["contador_id"] or creados_personal.get(d["contador_email"])
        cliente = Cliente(
            nombre_comercial=d["nombre_comercial"], razon_social=d["razon_social"],
            rfc=d["rfc"], telefono_whatsapp=d["telefono_whatsapp"],
            email=d["email"], tipo_cliente=TIPOS_CLIENTE[d["tipo_cliente"]],
            tipo_persona=d["tipo_persona"], regimen_fiscal=d["regimen_fiscal"],
            contador_asignado_id=contador_id,
            tiene_imss=d["tiene_imss"], tiene_nomina=d["tiene_nomina"],
            periodicidad_nomina=d["periodicidad_nomina"],
            honorario_mensual=d["honorario_mensual"],
            periodicidad_honorario=d["periodicidad_honorario"],
            dia_corte_honorario=d["dia_corte_honorario"],
            bd_contpaq_contabilidad=d["bd_contpaq_contabilidad"],
            bd_contpaq_nomina=d["bd_contpaq_nomina"])
        db.add(cliente)
        db.flush()

        # Lo que ya debía antes del alta
        if d["adeudo_previo_monto"]:
            from models.models import AdeudoPrevio
            db.add(AdeudoPrevio(
                cliente_id=cliente.id,
                concepto=(d["adeudo_previo_concepto"]
                          or "Saldo anterior al alta en el sistema")[:200],
                monto_original=d["adeudo_previo_monto"], monto_pagado=0,
                registrado_por_id=u.id))
            db.flush()

        if d["crear_portal"]:
            temporal = generar_password_temporal()
            portal = Usuario(nombre=d["nombre_comercial"], email=d["email"],
                             rol=RolUsuario.CLIENTE,
                             password_hash=hash_password(temporal),
                             debe_cambiar_password=True)
            db.add(portal)
            db.flush()
            cliente.usuario_portal_id = portal.id
            credenciales.append({"tipo": "Cliente", "nombre": d["nombre_comercial"],
                                 "usuario": d["email"], "password": temporal,
                                 "rol": "cliente"})

    auditoria.registrar(db, usuario_id=u.id, accion="alta_masiva",
                        tabla_afectada="clientes", registro_id=None,
                        request=request,
                        personal_creado=len(personal),
                        clientes_creados=len(clientes),
                        archivo=archivo.filename)
    db.commit()
    return {"ok": True,
            "personal_creado": len(personal),
            "clientes_creados": len(clientes),
            "credenciales": credenciales,
            "aviso": ("Anote estas contraseñas: NO se vuelven a mostrar. "
                      "Cada quien deberá cambiarla en su primer acceso.")}
